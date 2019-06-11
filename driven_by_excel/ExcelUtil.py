#!/usr/bin/python
# -*- encoding: utf-8 -*-

from openpyxl import *
import os,time
from ProjVar import *

class ParseExcel(object):

    def __init__(self,excelFilePath):
        if not os.path.exists(excelFilePath):
            self.wb = None
        self.excelFilePath = excelFilePath
        self.wb = load_workbook(self.excelFilePath)
        self.ws = self.set_sheet_by_name(self.wb.sheetnames[0])

    def get_col_value(self,col_no):
        # 第一步：根据列号获取一列
        # 第二步：遍历这列中的每个元素，并获取元素的值
        data = []
        if not isinstance(col_no, int):
            return None
        try:
            for cell in list(self.ws.columns)[col_no-1]:
                data.append(cell.value)
        except:
            return None
        else:
            return data[1:]

    def write_cell(self, row_no, col_no, content):
        """参数行号和列表从1开始表示第一行"""
        if (not isinstance(row_no, int)) or (not isinstance(col_no, int)):
            return None
        try:
            self.ws.cell(row=row_no, column=col_no).value = content
            self.wb.save(self.excelFilePath)
        except:
            return None
        self._save()

    def write_exec_time(self, row_no, col_no):
        timeTup = time.localtime()
        currentDate = str(timeTup.tm_year) + "年" + \
                      str(timeTup.tm_mon) + "月" + str(timeTup.tm_mday) + "日"
        currentTime = str(timeTup.tm_hour) + "时" + \
                      str(timeTup.tm_min) + "分" + str(timeTup.tm_sec) + "秒"
        self.write_cell(row_no, col_no, currentDate + " " + currentTime)
        self._save()

    def set_sheet_by_name(self,name):
        if name in self.wb.sheetnames:
            self.ws = self.wb[name]
            return self.ws
        self.ws = None
        return self.ws

    def _save(self):
        # 表格中写入数据，保存生效
        self.wb.save(self.excelFilePath)

if __name__ == "__main__":
    test_excel = ParseExcel("data.xlsx")
    test_excel.set_sheet_by_name("搜索数据表")
    test_data_list = test_excel.get_col_value(test_data_col_no)
    expect_data_list = test_excel.get_col_value(test_expect_data_col_no)
    data = list(zip(test_data_list, expect_data_list))
    print(data)